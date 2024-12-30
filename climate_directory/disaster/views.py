from django.shortcuts import render
from openpyxl import load_workbook


def get_disaster_data(request):
    """
    View to fetch and display disaster data based on user-selected filters.
    """
    data = None
    error = None
    disaster_types = []  # List to hold disaster types
    subregions = []  # List to hold subregions

    # Filepath and column definitions
    filename = r"D:\Natural_disaster\natural_disaster\Disaster_datas.xlsx"
    disaster_type_column = 'Disaster Type'
    country_column = 'Country'
    disno_column = 'DisNo.'
    subregion_column = 'Subregion'
    exclude_columns = ['Admin Unit']  # Columns to exclude from the output

    try:
        # Load the workbook and get the active sheet
        book = load_workbook(filename, data_only=True)
        sheet = book.active

        # Extract headers from the first row
        headers = [cell.value for cell in sheet[1]]

        # Extract unique disaster types
        if disaster_type_column in headers:
            disaster_type_index = headers.index(disaster_type_column)
            disaster_types = sorted(
                set(row[disaster_type_index] for row in sheet.iter_rows(min_row=2, values_only=True) if row[disaster_type_index])
            )

        # Extract unique subregions
        if subregion_column in headers:
            subregion_index = headers.index(subregion_column)
            subregions = sorted(
                set(row[subregion_index] for row in sheet.iter_rows(min_row=2, values_only=True) if row[subregion_index])
            )

    except FileNotFoundError:
        error = f"File not found: {filename}. Please ensure the file path is correct."
        return render(request, 'get_disaster_data.html', {'error': error})
    except Exception as e:
        error = f"An unexpected error occurred: {str(e)}"
        return render(request, 'get_disaster_data.html', {'error': error})

    if request.method == "POST":
        # Extract user inputs
        country_name = request.POST.get('country', '').strip()
        subregion_name = request.POST.get('subregion', '').strip()
        start_year = request.POST.get('start_year', '').strip()
        end_year = request.POST.get('end_year', '').strip()
        disaster_type = request.POST.get('disaster_type', '').strip()

        try:
            results = []

            # Search for rows matching the filters
            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_data = dict(zip(headers, row))

                # Extract the year from the "DisNo." column
                disno = row_data.get(disno_column, '')
                year = disno.split('-')[0] if disno else None

                # Apply filters
                if (
                    (not country_name or row_data.get(country_column) == country_name) and
                    (not subregion_name or row_data.get(subregion_column) == subregion_name) and
                    (not start_year or (year and year >= start_year)) and
                    (not end_year or (year and year <= end_year)) and
                    (not disaster_type or row_data.get(disaster_type_column) == disaster_type)
                ):
                    # Remove excluded columns
                    for col in exclude_columns:
                        row_data.pop(col, None)
                    results.append(row_data)

            if results:
                # Sort results by year (descending)
                results.sort(key=lambda x: x.get(disno_column, ''), reverse=True)
                data = results
            else:
                error = (
                    "No disaster data found for the specified filters. "
                    "Please adjust your criteria and try again."
                )

        except Exception as e:
            error = f"An unexpected error occurred while filtering data: {str(e)}"

    return render(
        request,
        'get_disaster_data.html',
        {
            'data': data,
            'error': error,
            'disaster_types': disaster_types,
            'subregions': subregions,
        },
    )
